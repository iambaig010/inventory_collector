# tests/test_excel_handler.py
import pytest
import tempfile
import os
from openpyxl import load_workbook
from src.utils.excel_handler import ExcelHandler

class TestExcelHandler:
    @pytest.fixture
    def excel_handler(self):
        return ExcelHandler()
        
    @pytest.fixture
    def sample_results(self):
        return [
            {
                'device_info': {
                    'hostname': 'test-switch-1',
                    'ip_address': '192.168.1.1',
                    'device_type': 'cisco_ios',
                    'connection_status': 'success'
                },
                'parsed_data': {
                    'model': 'WS-C2960X-24TS-L',
                    'serial_number': 'FOC1932X0K1',
                    'version': '15.2(4)E10',
                    'uptime': '1 week, 2 days',
                    'interfaces': [{'name': 'Gi0/1'}, {'name': 'Gi0/2'}]
                },
                'errors': [],
                'collection_time': '2025-01-15T10:30:00'
            },
            {
                'device_info': {
                    'hostname': 'test-switch-2',
                    'ip_address': '192.168.1.2',
                    'device_type': 'cisco_ios',
                    'connection_status': 'failed'
                },
                'parsed_data': {
                    'model': 'Unknown',
                    'serial_number': 'Unknown',
                    'version': 'Unknown',
                    'uptime': 'Unknown',
                    'interfaces': []
                },
                'errors': ['Connection timeout'],
                'collection_time': '2025-01-15T10:30:00'
            }
        ]
        
    @pytest.fixture
    def sample_summary(self):
        return {
            'total_devices': 2,
            'successful': 1,
            'failed': 1,
            'success_rate': 50.0,
            'device_types': {'cisco_ios': 2},
            'status_counts': {'success': 1, 'failed': 1},
            'collection_time': '2025-01-15T10:30:00'
        }
        
    def test_generate_report(self, excel_handler, sample_results, sample_summary):
        """Test complete report generation"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            excel_handler.generate_report(sample_results, tmp.name, sample_summary)
            
            # Verify file was created
            assert os.path.exists(tmp.name)
            
            # Load and verify content
            wb = load_workbook(tmp.name)
            
            # Check sheet names
            expected_sheets = ['Device Inventory', 'Summary', 'Errors']
            assert all(sheet in wb.sheetnames for sheet in expected_sheets)
            
            # Check inventory data
            inventory_ws = wb['Device Inventory']
            assert inventory_ws['A1'].value == 'Hostname'
            assert inventory_ws['A2'].value == 'test-switch-1'
            assert inventory_ws['H2'].value == 'success'  # Connection Status
            
            # Check summary data
            summary_ws = wb['Summary']
            assert 'Collection Summary' in [cell.value for row in summary_ws.iter_rows() for cell in row]
            
            # Check error data
            error_ws = wb['Errors']
            assert error_ws['A1'].value == 'Hostname'
            
        # Cleanup
        os.unlink(tmp.name)
        
    def test_create_device_template(self, excel_handler):
        """Test device template creation"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            excel_handler.create_device_template(tmp.name)
            
            # Verify file was created
            assert os.path.exists(tmp.name)
            
            # Load and verify content
            wb = load_workbook(tmp.name)
            
            # Check sheet names
            assert 'Device List' in wb.sheetnames
            assert 'Instructions' in wb.sheetnames
            
            # Check device list structure
            device_ws = wb['Device List']
            assert device_ws['A1'].value == 'hostname'
            assert device_ws['B1'].value == 'ip_address'
            
            # Check sample data
            assert device_ws['A2'].value == 'CORP-CORE-SW01'
            
        # Cleanup
        os.unlink(tmp.name)