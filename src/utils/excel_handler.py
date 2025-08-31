#!/usr/bin/env python3
"""
Enhanced Excel Handler - Prioritizes parsed hostnames in final report
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import logging
from typing import List, Dict, Any
from datetime import datetime

class ExcelHandler:
    """Enhanced Excel handler for IP-only input with hostname discovery"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def generate_report(self, inventory_results: List[Dict[str, Any]], 
                        output_path: str, summary: Dict[str, Any] = None):
        """Generate comprehensive Excel report with discovered hostnames"""
        self.logger.info(f"Generating Excel report: {output_path}")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default empty sheet
        
        # Main sheets
        self.create_main_inventory_sheet(wb, inventory_results)
        self.create_hostname_discovery_sheet(wb, inventory_results)
        self.create_interface_details_sheet(wb, inventory_results)
        if summary:
            self.create_summary_sheet(wb, summary)
        self.create_error_sheet(wb, inventory_results)
        
        # Save with timestamp in filename if not already present
        if not any(char.isdigit() for char in output_path):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_path = output_path.rsplit('.', 1)[0]
            extension = output_path.rsplit('.', 1)[1] if '.' in output_path else 'xlsx'
            output_path = f"{base_path}_{timestamp}.{extension}"
        
        wb.save(output_path)
        self.logger.info(f"Excel report saved: {output_path}")
        return output_path
        
    def create_main_inventory_sheet(self, wb: Workbook, results: List[Dict[str, Any]]):
        """Create main inventory sheet with discovered hostnames as primary identifier"""
        ws = wb.create_sheet("Device Inventory", 0)
        
        rows = []
        for result in results:
            device_info = result['device_info']
            parsed_data = result['parsed_data']
            
            # Prioritize discovered hostname over IP-based fallback
            discovered_hostname = parsed_data.get('hostname', '')
            display_hostname = discovered_hostname if discovered_hostname and not discovered_hostname.startswith('device-') else device_info.get('hostname', 'Unknown')
            
            # Determine if hostname was discovered or is fallback
            hostname_source = "Discovered" if discovered_hostname and not discovered_hostname.startswith('device-') else "Fallback"
            
            stack_info = parsed_data.get('stack_details', [])
            stack_count = len(stack_info) if stack_info else 1
            
            # Connection status with color coding info
            connection_status = device_info.get('connection_status', 'Unknown')
            status_indicator = "✓" if connection_status == 'success' else "✗"
            
            row = {
                'Status': status_indicator,
                'Hostname': display_hostname,
                'Hostname Source': hostname_source,
                'IP Address': device_info.get('ip_address', 'Unknown'),
                'Device Type': device_info.get('device_type', 'Unknown'),
                'Model': parsed_data.get('model', 'Unknown'),
                'Chassis Serial': parsed_data.get('serial_number', 'Unknown'),
                'System Serial': parsed_data.get('system_serial', parsed_data.get('serial_number', 'Unknown')),
                'Software Version': parsed_data.get('version', 'Unknown'),
                'IOS Version': parsed_data.get('ios_version', parsed_data.get('version', 'Unknown')),
                'Base MAC': parsed_data.get('base_mac', 'Unknown'),
                'Uptime': parsed_data.get('uptime', 'Unknown'),
                'Stack Count': stack_count,
                'Module Count': parsed_data.get('module_count', 0),
                'Hardware Modules': parsed_data.get('total_modules', 0),
                'Interface Count': len(parsed_data.get('interfaces', [])),
                'Connection Status': connection_status,
                'Has Errors': 'Yes' if result.get('errors') else 'No',
                'Collection Time': result.get('collection_time', 'Unknown'),
                'Location': device_info.get('location', ''),
                'Description': device_info.get('description', '')
            }
            rows.append(row)
            
        df = pd.DataFrame(rows)
        
        # Write data to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        # Format the sheet
        self.format_main_sheet(ws, len(rows))
        
    def create_hostname_discovery_sheet(self, wb: Workbook, results: List[Dict[str, Any]]):
        """Create sheet showing hostname discovery details"""
        ws = wb.create_sheet("Hostname Discovery")
        
        # Headers
        headers = ["IP Address", "Discovered Hostname", "Source", "Discovery Method", "Original Hostname", "Status"]
        ws.append(headers)
        
        # Data rows
        for result in results:
            device_info = result['device_info']
            parsed_data = result['parsed_data']
            
            ip_address = device_info.get('ip_address', 'Unknown')
            discovered_hostname = parsed_data.get('hostname', '')
            original_hostname = f"device-{ip_address.replace('.', '-')}"
            
            # Determine discovery details
            if discovered_hostname and not discovered_hostname.startswith('device-'):
                source = "Device Parsing"
                method = "CLI Command Output"
                status = "✓ Successfully Discovered"
            else:
                source = "Fallback"
                method = "IP-based Generation"
                status = "⚠ Not Discovered"
                discovered_hostname = original_hostname
            
            ws.append([
                ip_address,
                discovered_hostname,
                source,
                method,
                original_hostname,
                status
            ])
        
        # Format headers
        self.format_headers(ws)
        self.auto_adjust_columns(ws)

    def create_interface_details_sheet(self, wb: Workbook, results: List[Dict[str, Any]]):
        """Create detailed interface information sheet"""
        ws = wb.create_sheet("Interface Details")
        
        # Headers
        headers = ["Hostname", "IP Address", "Interface", "Status", "Protocol", "IP", "Description", "Speed", "Duplex"]
        ws.append(headers)
        
        # Data rows
        for result in results:
            device_info = result['device_info']
            parsed_data = result['parsed_data']
            
            hostname = device_info.get('hostname', 'Unknown')
            ip_address = device_info.get('ip_address', 'Unknown')
            
            interfaces = parsed_data.get('interfaces', [])
            if not interfaces:
                # Add a row even if no interfaces found
                ws.append([hostname, ip_address, 'No interfaces found', '', '', '', '', '', ''])
            else:
                for interface in interfaces:
                    ws.append([
                        hostname,
                        ip_address,
                        interface.get('name', 'Unknown'),
                        interface.get('status', 'Unknown'),
                        interface.get('protocol', 'Unknown'),
                        interface.get('ip', 'Unknown'),
                        interface.get('description', ''),
                        interface.get('speed', 'Unknown'),
                        interface.get('duplex', 'Unknown')
                    ])
        
        # Format headers
        self.format_headers(ws)
        self.auto_adjust_columns(ws)

    def format_main_sheet(self, ws, data_rows: int):
        """Format the main inventory sheet with enhanced styling"""
        # Headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Status column conditional formatting (Column A)
        for row in range(2, data_rows + 2):
            status_cell = ws[f'A{row}']
            if status_cell.value == "✓":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status_cell.value == "✗":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Hostname source highlighting (Column C)
        for row in range(2, data_rows + 2):
            source_cell = ws[f'C{row}']
            if source_cell.value == "Discovered":
                source_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif source_cell.value == "Fallback":
                source_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=data_rows + 1):
            for cell in row:
                cell.border = thin_border

    def format_headers(self, ws):
        """Format sheet headers with consistent styling"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set column width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_summary_sheet(self, wb: Workbook, summary: Dict[str, Any]):
        """Create enhanced summary sheet with hostname discovery stats"""
        ws = wb.create_sheet("Collection Summary")
        
        # Title
        ws.append(["Network Device Inventory Collection Summary"])
        ws.append([])  # Empty row
        
        # Basic stats
        ws.append(["Basic Statistics", ""])
        ws.append(["Total Devices", summary.get('total_devices', 0)])
        ws.append(["Successful Connections", summary.get('successful', 0)])
        ws.append(["Failed Connections", summary.get('failed', 0)])
        ws.append(["Success Rate", f"{summary.get('success_rate', 0):.1f}%"])
        ws.append([])  # Empty row
        
        # Hostname discovery stats
        ws.append(["Hostname Discovery", ""])
        ws.append(["Hostnames Discovered", len(summary.get('discovered_hostnames', []))])
        ws.append(["Discovery Rate", f"{summary.get('hostname_discovery_rate', 0):.1f}%"])
        ws.append([])  # Empty row
        
        # Device type breakdown
        ws.append(["Device Types", "Count"])
        device_types = summary.get('device_types', {})
        for device_type, count in device_types.items():
            ws.append([device_type, count])
        ws.append([])  # Empty row
        
        # Connection status breakdown
        ws.append(["Connection Status", "Count"])
        status_counts = summary.get('status_counts', {})
        for status, count in status_counts.items():
            ws.append([status, count])
        ws.append([])  # Empty row
        
        # Discovered hostnames list
        discovered_hostnames = summary.get('discovered_hostnames', [])
        if discovered_hostnames:
            ws.append(["Successfully Discovered Hostnames", ""])
            for hostname in discovered_hostnames:
                ws.append([hostname, ""])
        
        # Collection info
        ws.append([])
        ws.append(["Collection Time", summary.get('collection_time', 'Unknown')])
        
        # Format the summary sheet
        self.format_summary_sheet(ws)

    def format_summary_sheet(self, ws):
        """Format the summary sheet with enhanced styling"""
        # Title formatting
        title_font = Font(size=16, bold=True, color="2F5597")
        ws['A1'].font = title_font
        
        # Section headers
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        section_font = Font(bold=True)
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Section headers
                    if cell.value in ["Basic Statistics", "Hostname Discovery", "Device Types", 
                                    "Connection Status", "Successfully Discovered Hostnames"]:
                        cell.font = section_font
                        cell.fill = section_fill
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)

    def create_error_sheet(self, wb: Workbook, results: List[Dict[str, Any]]):
        """Create detailed error analysis sheet"""
        ws = wb.create_sheet("Error Analysis")
        
        # Headers
        headers = ["Hostname", "IP Address", "Connection Status", "Error Category", "Error Details", "Timestamp"]
        ws.append(headers)
        
        has_errors = False
        error_categories = {
            'connection': ['timeout', 'connection refused', 'unreachable'],
            'authentication': ['authentication', 'login', 'permission denied'],
            'command': ['command not found', 'invalid command', 'syntax error'],
            'parsing': ['parse', 'format', 'unexpected output']
        }
        
        for result in results:
            device_info = result['device_info']
            hostname = result['parsed_data'].get('hostname', device_info.get('hostname', 'Unknown'))
            ip_address = device_info.get('ip_address', 'Unknown')
            connection_status = device_info.get('connection_status', 'Unknown')
            timestamp = result.get('collection_time', 'Unknown')
            
            # Process errors
            if errors := result.get('errors'):
                has_errors = True
                for err in errors:
                    error_msg = str(err).lower()
                    
                    # Categorize error
                    error_category = 'Other'
                    for category, keywords in error_categories.items():
                        if any(keyword in error_msg for keyword in keywords):
                            error_category = category.title()
                            break
                    
                    ws.append([hostname, ip_address, connection_status, error_category, str(err), timestamp])
            
            # Also add connection failures even without explicit errors
            elif connection_status == 'failed':
                has_errors = True
                ws.append([hostname, ip_address, connection_status, 'Connection', 'Connection failed', timestamp])
        
        # If no errors found, add a note
        if not has_errors:
            ws.append(["✓ Success", "All devices processed successfully", "", "No Errors", "", datetime.now().isoformat()])
        
        # Format headers and apply conditional formatting
        self.format_headers(ws)
        
        # Color code error categories
        category_colors = {
            'Connection': 'FFCDD2',
            'Authentication': 'FFE0B2', 
            'Command': 'F8BBD9',
            'Parsing': 'D1C4E9',
            'Other': 'CFD8DC'
        }
        
        for row in range(2, ws.max_row + 1):
            category_cell = ws[f'D{row}']
            if category_cell.value in category_colors:
                fill_color = category_colors[category_cell.value]
                for col in range(1, 7):  # Apply to entire row
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
        
        self.auto_adjust_columns(ws)

    def create_device_template(self, output_file: str):
        """Create simplified Excel template for IP addresses only"""
        try:
            # Simple template with just IP addresses and optional fields
            template_data = {
                'ip_address': [
                    '192.168.1.10',
                    '192.168.1.11', 
                    '192.168.1.12',
                    '10.0.0.1',
                    '10.0.0.2'
                ],
                'device_type': [
                    'cisco_ios',
                    'cisco_ios',
                    'hirschmann_hios',
                    'cisco_xe',
                    'juniper_junos'
                ],
                'port': [22, 22, 22, 22, 22],
                'location': [
                    'Server Room A',
                    'Server Room A', 
                    'Floor 2 Closet',
                    'Core Network',
                    'Distribution'
                ],
                'description': [
                    'Access Switch 1',
                    'Access Switch 2',
                    'Industrial Switch',
                    'Core Router',
                    'Distribution Switch'
                ]
            }
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Device List"
            
            # Create the main template
            df = pd.DataFrame(template_data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            self.format_headers(ws)
            self.auto_adjust_columns(ws)
            
            # Create instructions sheet
            instructions_ws = wb.create_sheet("Instructions")
            instructions_data = [
                ["Network Device Inventory - IP Address Template"],
                [""],
                ["IMPORTANT: Only IP addresses are required in this Excel file!"],
                [""],
                ["Instructions:"],
                ["1. Fill in the 'ip_address' column with device IP addresses"],
                ["2. Username and password will be entered in the GUI application"],
                ["3. Device type can be set per device or globally in GUI"],
                ["4. Port, location, and description are optional"],
                ["5. Hostnames will be automatically discovered from devices"],
                [""],
                ["Required Column:"],
                ["• ip_address - The IP address of the network device"],
                [""],
                ["Optional Columns:"],
                ["• device_type - Override default device type (cisco_ios, cisco_xe, etc.)"],
                ["• port - SSH port (default: 22)"],
                ["• location - Physical location of device"],
                ["• description - Device description/notes"],
                [""],
                ["Supported Device Types:"],
                ["• cisco_ios - Cisco IOS devices"],
                ["• cisco_xe - Cisco IOS-XE devices"],
                ["• cisco_nxos - Cisco Nexus devices"],
                ["• juniper_junos - Juniper JunOS devices"],
                ["• hirschmann_hios - Hirschmann HiOS devices"],
                ["• generic_ssh - Generic SSH devices"],
                [""],
                ["Output:"],
                ["The final report will contain discovered hostnames as the primary"],
                ["identifier, extracted directly from the device configurations."]
            ]
            
            for row in instructions_data:
                instructions_ws.append(row)
            
            # Format instructions
            instructions_ws['A1'].font = Font(size=14, bold=True, color="2F5597")
            
            # Highlight important notes
            for row in range(3, 6):
                cell = instructions_ws[f'A{row}']
                if cell.value and "IMPORTANT" in str(cell.value):
                    cell.font = Font(bold=True, color="D63384")
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            self.auto_adjust_columns(instructions_ws)
            
            wb.save(output_file)
            self.logger.info(f"IP-only device template created: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating template: {e}")
            raise

    def export_hostnames_csv(self, results: List[Dict[str, Any]], output_file: str):
        """Export discovered hostnames to CSV for reference"""
        try:
            hostname_data = []
            
            for result in results:
                device_info = result['device_info']
                parsed_data = result['parsed_data']
                
                ip_address = device_info.get('ip_address', 'Unknown')
                discovered_hostname = parsed_data.get('hostname', '')
                connection_status = device_info.get('connection_status', 'Unknown')
                
                # Determine if hostname was actually discovered
                is_discovered = discovered_hostname and not discovered_hostname.startswith('device-')
                
                hostname_data.append({
                    'IP_Address': ip_address,
                    'Discovered_Hostname': discovered_hostname if is_discovered else '',
                    'Fallback_Hostname': f"device-{ip_address.replace('.', '-')}",
                    'Final_Hostname': discovered_hostname if discovered_hostname else f"device-{ip_address.replace('.', '-')}",
                    'Discovery_Status': 'Discovered' if is_discovered else 'Fallback',
                    'Connection_Status': connection_status,
                    'Device_Type': device_info.get('device_type', 'Unknown')
                })
            
            df = pd.DataFrame(hostname_data)
            df.to_csv(output_file, index=False)
            
            self.logger.info(f"Hostname discovery report exported to CSV: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error exporting hostnames CSV: {e}")
            raise

    def generate_comparison_report(self, results: List[Dict[str, Any]], 
                                 previous_results: List[Dict[str, Any]], 
                                 output_path: str):
        """Generate comparison report between current and previous inventory"""
        self.logger.info("Generating comparison report...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create comparison sheet
        ws = wb.create_sheet("Inventory Comparison", 0)
        
        # Headers
        headers = ["IP Address", "Current Hostname", "Previous Hostname", "Status Change", 
                  "Current Model", "Previous Model", "Current Version", "Previous Version", "Notes"]
        ws.append(headers)
        
        # Create lookup dictionaries
        current_devices = {r['device_info']['ip_address']: r for r in results}
        previous_devices = {r['device_info']['ip_address']: r for r in previous_results}
        
        all_ips = set(current_devices.keys()) | set(previous_devices.keys())
        
        for ip in sorted(all_ips):
            current = current_devices.get(ip)
            previous = previous_devices.get(ip)
            
            if current and previous:
                # Device exists in both
                curr_hostname = current['parsed_data'].get('hostname', 'Unknown')
                prev_hostname = previous['parsed_data'].get('hostname', 'Unknown')
                
                status_change = "No Change"
                if curr_hostname != prev_hostname:
                    status_change = "Hostname Changed"
                
                notes = []
                if current['parsed_data'].get('version') != previous['parsed_data'].get('version'):
                    notes.append("Version Changed")
                if current['parsed_data'].get('model') != previous['parsed_data'].get('model'):
                    notes.append("Model Changed")
                
                ws.append([
                    ip,
                    curr_hostname,
                    prev_hostname,
                    status_change,
                    current['parsed_data'].get('model', 'Unknown'),
                    previous['parsed_data'].get('model', 'Unknown'),
                    current['parsed_data'].get('version', 'Unknown'),
                    previous['parsed_data'].get('version', 'Unknown'),
                    "; ".join(notes) if notes else "No changes"
                ])
                
            elif current:
                # New device
                ws.append([
                    ip,
                    current['parsed_data'].get('hostname', 'Unknown'),
                    "N/A",
                    "New Device",
                    current['parsed_data'].get('model', 'Unknown'),
                    "N/A",
                    current['parsed_data'].get('version', 'Unknown'),
                    "N/A",
                    "Newly discovered device"
                ])
                
            elif previous:
                # Removed device
                ws.append([
                    ip,
                    "N/A",
                    previous['parsed_data'].get('hostname', 'Unknown'),
                    "Device Removed",
                    "N/A",
                    previous['parsed_data'].get('model', 'Unknown'),
                    "N/A",
                    previous['parsed_data'].get('version', 'Unknown'),
                    "Device no longer accessible"
                ])
        
        # Format comparison sheet
        self.format_headers(ws)
        
        # Color code status changes
        for row in range(2, ws.max_row + 1):
            status_cell = ws[f'D{row}']
            if status_cell.value == "New Device":
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                    )
            elif status_cell.value == "Device Removed":
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
                    )
            elif status_cell.value == "Hostname Changed":
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
        
        self.auto_adjust_columns(ws)
        wb.save(output_path)
        self.logger.info(f"Comparison report saved: {output_path}")